from urllib.request import urlretrieve
from openpyxl import load_workbook
import json
import re

table_url = "https://www.alko.fi/INTERSHOP/static/WFS/Alko-OnlineShop-Site/-/Alko-OnlineShop/fi_FI/Alkon%20Hinnasto%20Tekstitiedostona/alkon-hinnasto-tekstitiedostona.xlsx"
filename_in = "hinnasto.xlsx"
filename_out = "drinkdata.json"

data = {}
data["drinks"] = []
data["nondrinks"] = []

print("Downloading file")
urlretrieve(table_url, filename_in)
print("Download ready")

wb = load_workbook(filename_in, data_only=True)
ws = wb.worksheets[0] #tämä lataa ensimmäisen sivun excel-tiedostosta

def compute_img_url(name, id):
    name = re.sub(r'([^\s\w]|_)+', '', name)
    name = name.replace("ä", "a")
    name = name.replace("ö", "o")
    name = name.replace("å", "a")
    name = name.replace(" ", "-")

    url = "https://images.alko.fi/images/cs_srgb,f_auto,t_products/cdn/"+str(id)+"/"+name+".jpg"
    return url

drink_calc = 0
non_drink_calc = 0
row_count = ws.max_row-4

for row in ws.iter_rows(min_row=5, max_col=22, values_only=True):
    id = row[0]
    name = row[1]
    price = float(row[4])
    url = "https://alko.fi/tuotteet/"+id
    imgUrl = compute_img_url(name, id)

    if row[3]: #jos tilavuus on määritelty
        manufacturer = row[2]
        volume = float(row[3].strip(" l").replace(",","."))
        type = row[8]
        alcohol = float(row[21])

        ppL = 0
        ppLe = "approaches infinity"

        if volume != 0:
            ppL = round(price/volume, 2)
            if alcohol != 0:
                ppLe = round(price*100/(volume*alcohol), 2)

        drink = { 
            "id": id,
            "name": name,
            "manufacturer": manufacturer,
            "size": volume,
            "price": price,
            "type": type,
            "alcohol": alcohol,
            "priceperL": ppL,
            "priceperethanolL": ppLe,
            "url": url,
            "imgUrl": imgUrl
        }

        data["drinks"].append(drink)
        drink_calc += 1

    else:
        nondrink = {
            "id": id,
            "name": name,
            "price": price,
            "url": url,
            "imgUrl": imgUrl,
        }
        data["nondrinks"].append(nondrink)
        non_drink_calc += 1

    print(str(drink_calc+non_drink_calc)+"/"+str(row_count))

with open(filename_out, "w") as outfile:
    json.dump(data, outfile)

print(str(drink_calc)+" drinks found")
print(str(non_drink_calc)+" non-drinks found")